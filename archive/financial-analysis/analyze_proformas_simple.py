#!/usr/bin/env python3
"""
Simple Financial ProForma Analysis Tool
Analyzes Excel proforma templates without Unicode issues
"""

import pandas as pd
import openpyxl
import os
import json

def analyze_excel_file(file_path):
    """Analyze a single Excel proforma file."""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {os.path.basename(file_path)}")
    print(f"{'='*80}")

    try:
        # Load workbook with formulas preserved
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)

        analysis = {
            'file_name': os.path.basename(file_path),
            'worksheets': [],
            'key_calculations': {}
        }

        for sheet_name in workbook.sheetnames:
            print(f"\nWORKSHEET: {sheet_name}")
            print("-" * 50)

            ws_formula = workbook[sheet_name]
            ws_data = workbook_data[sheet_name]

            # Extract key financial data and formulas
            formulas = []
            financial_data = {}

            for row in range(1, min(ws_formula.max_row + 1, 100)):
                for col in range(1, min(ws_formula.max_column + 1, 15)):
                    cell_formula = ws_formula.cell(row=row, column=col).value
                    cell_data = ws_data.cell(row=row, column=col).value

                    # Look for formulas
                    if cell_formula and str(cell_formula).startswith('='):
                        cell_ref = f"{chr(64+col)}{row}"
                        formulas.append({
                            'cell': cell_ref,
                            'formula': str(cell_formula),
                            'result': cell_data
                        })

                    # Look for financial terms
                    if cell_data and isinstance(cell_data, str):
                        cell_lower = cell_data.lower()
                        financial_terms = [
                            'gross income', 'net income', 'cash flow', 'cap rate',
                            'vacancy', 'expenses', 'debt service', 'roi', 'irr',
                            'purchase price', 'down payment', 'loan amount',
                            'rental income', 'operating expenses', 'insurance',
                            'property tax', 'maintenance', 'management fee'
                        ]

                        for term in financial_terms:
                            if term in cell_lower:
                                cell_ref = f"{chr(64+col)}{row}"
                                adjacent_value = ws_data.cell(row=row, column=col+1).value
                                if adjacent_value and isinstance(adjacent_value, (int, float)):
                                    financial_data[term] = {
                                        'cell': cell_ref,
                                        'value': adjacent_value,
                                        'label': cell_data
                                    }

            print(f"Found {len(formulas)} formulas")
            print(f"Found {len(financial_data)} financial data points")

            # Show key financial data
            for term, data in financial_data.items():
                print(f"  {term}: ${data['value']:,.2f} ({data['cell']})")

            analysis['worksheets'].append({
                'name': sheet_name,
                'formulas': formulas,
                'financial_data': financial_data
            })

        return analysis

    except Exception as e:
        print(f"ERROR: {e}")
        return None

def main():
    """Main function."""
    print("FINANCIAL PROFORMA ANALYSIS")
    print("="*80)

    proforma_files = [
        r"C:\Users\gsima\Documents\MyProject\ConcordBroker\TEMP\FINANCIAL ANALYSIS\Greenwich Place ProForma 07 27 11.xlsx",
        r"C:\Users\gsima\Documents\MyProject\ConcordBroker\TEMP\FINANCIAL ANALYSIS\Proforma Oakland Park 05.21.24 Lauren Updated 08.19.24 JP (1).xlsx",
        r"C:\Users\gsima\Documents\MyProject\ConcordBroker\TEMP\FINANCIAL ANALYSIS\6201 TAYLOR STREET, HOLLYWOOD FL 33024.xlsx"
    ]

    analyses = []
    for file_path in proforma_files:
        if os.path.exists(file_path):
            analysis = analyze_excel_file(file_path)
            if analysis:
                analyses.append(analysis)
        else:
            print(f"File not found: {file_path}")

    # Compile ultimate structure
    print(f"\n{'='*80}")
    print("ULTIMATE INVESTMENT ANALYSIS STRUCTURE")
    print(f"{'='*80}")

    ultimate_structure = {
        'acquisition': {
            'purchase_price': 'Property purchase price',
            'down_payment': 'Initial cash investment',
            'loan_amount': 'Financing amount',
            'closing_costs': 'Transaction costs'
        },
        'financing': {
            'interest_rate': 'Annual interest rate',
            'loan_term': 'Loan term in years',
            'monthly_payment': 'Principal + Interest payment',
            'annual_debt_service': 'Total annual debt payments'
        },
        'income': {
            'gross_rental_income': 'Total potential rental income',
            'vacancy_allowance': 'Vacancy and collection loss',
            'effective_gross_income': 'Gross income - vacancy',
            'other_income': 'Parking, laundry, etc.'
        },
        'expenses': {
            'property_taxes': 'Annual property taxes',
            'insurance': 'Property insurance premiums',
            'property_management': 'Management fees',
            'maintenance_repairs': 'Property maintenance costs',
            'utilities': 'Landlord-paid utilities',
            'other_expenses': 'Miscellaneous operating expenses'
        },
        'cash_flow': {
            'net_operating_income': 'Income - Operating Expenses',
            'before_tax_cash_flow': 'NOI - Debt Service',
            'after_tax_cash_flow': 'After tax considerations',
            'annual_cash_flow': 'Total annual cash generation'
        },
        'returns': {
            'cap_rate': 'NOI / Purchase Price',
            'cash_on_cash_return': 'Cash Flow / Cash Invested',
            'total_roi': 'Total return including appreciation',
            'internal_rate_of_return': 'IRR over holding period'
        }
    }

    for category, items in ultimate_structure.items():
        print(f"\n{category.upper()}")
        for key, description in items.items():
            print(f"  {key.replace('_', ' ').title()}: {description}")

    # Save analysis
    with open('proforma_analysis.json', 'w') as f:
        json.dump({
            'analyses': analyses,
            'structure': ultimate_structure
        }, f, indent=2, default=str)

    print(f"\nAnalysis saved to proforma_analysis.json")

    return ultimate_structure

if __name__ == "__main__":
    main()