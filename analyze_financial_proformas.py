#!/usr/bin/env python3
"""
Financial ProForma Analysis Tool
Analyzes Excel proforma templates to extract structure, formulas, and calculations
for designing the ultimate Investment Analysis component.
"""

import pandas as pd
import openpyxl
import os
import json
from typing import Dict, List, Any

def analyze_excel_structure(file_path: str) -> Dict[str, Any]:
    """Analyze Excel file structure, formulas, and data."""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {os.path.basename(file_path)}")
    print(f"{'='*80}")

    try:
        # Load workbook with formulas preserved
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        workbook_data = openpyxl.load_workbook(file_path, data_only=True)

        analysis = {
            'file_name': os.path.basename(file_path),
            'worksheets': [],
            'key_formulas': [],
            'financial_metrics': {},
            'data_structure': {}
        }

        for sheet_name in workbook.sheetnames:
            print(f"\nWORKSHEET: {sheet_name}")
            print("-" * 50)

            ws_formula = workbook[sheet_name]
            ws_data = workbook_data[sheet_name]

            sheet_analysis = {
                'name': sheet_name,
                'dimensions': f"{ws_formula.max_row}x{ws_formula.max_column}",
                'formulas': [],
                'key_data': {},
                'structure': []
            }

            # Extract all formulas and their locations
            formula_count = 0
            for row in range(1, min(ws_formula.max_row + 1, 100)):  # Limit to first 100 rows
                for col in range(1, min(ws_formula.max_column + 1, 20)):  # Limit to first 20 columns
                    cell_formula = ws_formula.cell(row=row, column=col).value
                    cell_data = ws_data.cell(row=row, column=col).value

                    if cell_formula and str(cell_formula).startswith('='):
                        formula_count += 1
                        cell_ref = f"{chr(64+col)}{row}"
                        formula_info = {
                            'cell': cell_ref,
                            'formula': str(cell_formula),
                            'result': cell_data
                        }
                        sheet_analysis['formulas'].append(formula_info)

                        # Print key formulas
                        if formula_count <= 10:  # Show first 10 formulas
                            print(f"  {cell_ref}: {cell_formula} = {cell_data}")

            # Extract key data patterns (look for financial terms)
            financial_terms = [
                'income', 'revenue', 'rent', 'gross', 'net', 'expense', 'cost',
                'cap rate', 'roi', 'irr', 'npv', 'cash flow', 'noi', 'vacancy',
                'appreciation', 'tax', 'insurance', 'maintenance', 'management',
                'debt service', 'loan', 'mortgage', 'interest', 'principal'
            ]

            for row in range(1, min(ws_formula.max_row + 1, 50)):
                for col in range(1, min(ws_formula.max_column + 1, 10)):
                    cell_value = ws_data.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        cell_lower = cell_value.lower()
                        for term in financial_terms:
                            if term in cell_lower:
                                cell_ref = f"{chr(64+col)}{row}"
                                if term not in sheet_analysis['key_data']:
                                    sheet_analysis['key_data'][term] = []
                                sheet_analysis['key_data'][term].append({
                                    'cell': cell_ref,
                                    'text': cell_value,
                                    'adjacent_value': ws_data.cell(row=row, column=col+1).value
                                })

            print(f"  📈 Found {formula_count} formulas")
            print(f"  🔍 Found {len(sheet_analysis['key_data'])} financial term categories")

            # Show key financial terms found
            for term, instances in sheet_analysis['key_data'].items():
                if instances:
                    print(f"    {term.upper()}: {len(instances)} instances")
                    for instance in instances[:3]:  # Show first 3 instances
                        print(f"      {instance['cell']}: {instance['text']} = {instance['adjacent_value']}")

            analysis['worksheets'].append(sheet_analysis)

        print(f"\n✅ ANALYSIS COMPLETE")
        print(f"   Worksheets: {len(analysis['worksheets'])}")
        print(f"   Total Formulas: {sum(len(ws['formulas']) for ws in analysis['worksheets'])}")

        return analysis

    except Exception as e:
        print(f"❌ ERROR analyzing {file_path}: {e}")
        return {'error': str(e)}

def extract_financial_calculations(analyses: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Extract common financial calculation patterns from all analyses."""
    print(f"\n{'='*80}")
    print("EXTRACTING FINANCIAL CALCULATION PATTERNS")
    print(f"{'='*80}")

    patterns = {
        'income_calculations': [],
        'expense_calculations': [],
        'cash_flow_calculations': [],
        'return_metrics': [],
        'debt_service': [],
        'common_formulas': []
    }

    # Analyze formulas across all files
    all_formulas = []
    for analysis in analyses:
        if 'error' not in analysis:
            for worksheet in analysis['worksheets']:
                all_formulas.extend(worksheet['formulas'])

    print(f"📊 Analyzing {len(all_formulas)} total formulas...")

    # Categorize formulas by financial function
    for formula_info in all_formulas:
        formula = formula_info['formula'].upper()

        # Income-related calculations
        if any(term in formula for term in ['RENT', 'INCOME', 'REVENUE', 'GROSS']):
            patterns['income_calculations'].append(formula_info)

        # Expense-related calculations
        elif any(term in formula for term in ['EXPENSE', 'COST', 'TAX', 'INSURANCE', 'MAINTENANCE']):
            patterns['expense_calculations'].append(formula_info)

        # Cash flow calculations
        elif any(term in formula for term in ['CASH', 'NET', 'NOI']):
            patterns['cash_flow_calculations'].append(formula_info)

        # Return metrics
        elif any(term in formula for term in ['RATE', 'ROI', 'IRR', 'NPV']):
            patterns['return_metrics'].append(formula_info)

        # Debt service
        elif any(term in formula for term in ['DEBT', 'LOAN', 'MORTGAGE', 'PAYMENT']):
            patterns['debt_service'].append(formula_info)

    # Show summary of patterns found
    for category, formulas in patterns.items():
        print(f"\n{category.upper().replace('_', ' ')}: {len(formulas)} formulas")
        for formula in formulas[:5]:  # Show first 5 examples
            print(f"  {formula['cell']}: {formula['formula'][:60]}{'...' if len(formula['formula']) > 60 else ''}")

    return patterns

def main():
    """Main analysis function."""
    proforma_files = [
        r"C:\Users\gsima\Documents\MyProject\ConcordBroker\TEMP\FINANCIAL ANALYSIS\Greenwich Place ProForma 07 27 11.xlsx",
        r"C:\Users\gsima\Documents\MyProject\ConcordBroker\TEMP\FINANCIAL ANALYSIS\Proforma Oakland Park 05.21.24 Lauren Updated 08.19.24 JP (1).xlsx",
        r"C:\Users\gsima\Documents\MyProject\ConcordBroker\TEMP\FINANCIAL ANALYSIS\6201 TAYLOR STREET, HOLLYWOOD FL 33024.xlsx"
    ]

    print("🏢 FINANCIAL PROFORMA ANALYSIS")
    print("="*80)
    print("Analyzing Excel proforma templates to design ultimate Investment Analysis component")

    analyses = []
    for file_path in proforma_files:
        if os.path.exists(file_path):
            analysis = analyze_excel_structure(file_path)
            analyses.append(analysis)
        else:
            print(f"❌ File not found: {file_path}")

    if analyses:
        # Extract common patterns
        patterns = extract_financial_calculations(analyses)

        # Generate ultimate proforma structure
        print(f"\n{'='*80}")
        print("ULTIMATE INVESTMENT ANALYSIS STRUCTURE")
        print(f"{'='*80}")

        ultimate_structure = {
            'sections': {
                'property_info': ['Purchase Price', 'Down Payment', 'Loan Amount', 'Interest Rate', 'Loan Term'],
                'income_analysis': ['Gross Rental Income', 'Vacancy Rate', 'Effective Gross Income', 'Other Income'],
                'expense_analysis': ['Property Taxes', 'Insurance', 'Property Management', 'Maintenance', 'Utilities', 'Other Expenses'],
                'cash_flow_analysis': ['Net Operating Income', 'Debt Service', 'Before Tax Cash Flow', 'After Tax Cash Flow'],
                'return_metrics': ['Cap Rate', 'Cash-on-Cash Return', 'ROI', 'IRR', 'NPV'],
                'sensitivity_analysis': ['Rent Growth', 'Expense Growth', 'Vacancy Scenarios', 'Interest Rate Changes']
            }
        }

        for section, items in ultimate_structure['sections'].items():
            print(f"\n📊 {section.upper().replace('_', ' ')}")
            for item in items:
                print(f"  • {item}")

        # Save analysis results
        with open('financial_analysis_report.json', 'w') as f:
            json.dump({
                'analyses': analyses,
                'patterns': patterns,
                'ultimate_structure': ultimate_structure
            }, f, indent=2, default=str)

        print(f"\n✅ Analysis complete! Report saved to financial_analysis_report.json")

        return ultimate_structure

    else:
        print("❌ No valid proforma files found for analysis")
        return None

if __name__ == "__main__":
    main()